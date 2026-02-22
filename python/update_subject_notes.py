#!/usr/bin/env python3
"""
A360 Option B: JSON-authoritative subject notes + optional Excel mirror.

Supports two invocation modes:

1) Update notes (used by AnchorPromptBuilder autosave):
   python update_subject_notes.py --subject S021 --notes "..." [--meta "{}"] [--subjectRoot ...] [--projectRoot ...] [--excel ...]

2) Create subject (used by SubjectCreate screen; legacy interface maintained):
   python update_subject_notes.py create-subject --sex Male --ethnicity Polynesian --fitz III --notes "..." [--subjectRoot ...] [--projectRoot ...] [--excel ...] [--timelineFolderName TimelineA]

Outputs a single JSON object to stdout. All diagnostics go to stderr.
"""

import argparse
import datetime as _dt
import json
import os
import re
import sys
from pathlib import Path
from typing import Any, Dict, Optional, Tuple


# ----------------------------
# helpers
# ----------------------------

def _utc_now_iso() -> str:
    return _dt.datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def _eprint(*args: Any) -> None:
    print(*args, file=sys.stderr)


def _safe_folder(name: str, fallback: str = "Unsorted") -> str:
    s = (name or "").strip()
    if not s:
        return fallback
    # Windows-illegal path chars
    s = re.sub(r'[<>:"/\\|?*]', "_", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s or fallback


def _normalize_sex(sex: str) -> str:
    s = (sex or "").strip()
    sl = s.lower()
    if sl.startswith("m"):
        return "Male"
    if sl.startswith("f"):
        return "Female"
    return _safe_folder(s, "Unsorted")


def _normalize_subject_id(subject: str) -> Tuple[str, int, str]:
    """
    Returns (subjectId like S021), numeric id (21), folder name like subject021
    """
    raw = (subject or "").strip()
    if not raw:
        raise ValueError("missing subject id")
    m = re.fullmatch(r"[sS]?(\d+)", raw)
    if not m:
        raise ValueError(f"invalid subject id: {raw}")
    n = int(m.group(1))
    # pad to 3 for backwards compatibility; allow >999
    width = max(3, len(str(n)))
    folder = f"subject{n:0{width}d}"
    sid = f"S{n:0{width}d}"
    return sid, n, folder


def _load_config(cwd: Path) -> Dict[str, Any]:
    """
    Best-effort load of a360.config.json.
    Priority:
      1) env A360_CONFIG_PATH
      2) cwd/a360.config.json
      3) cwd/../a360.config.json
    """
    candidates = []
    env = os.environ.get("A360_CONFIG_PATH")
    if env:
        candidates.append(Path(env))
    candidates.append(cwd / "a360.config.json")
    candidates.append(cwd.parent / "a360.config.json")

    for p in candidates:
        try:
            if p.is_file():
                return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            continue
    return {}


def _compute_aging_root(cfg: Dict[str, Any], subject_root_arg: Optional[str], project_root_arg: Optional[str], cwd: Path) -> Path:
    # Prefer explicit arg, then config keys
    base_str = subject_root_arg or cfg.get("subjectRoot") or cfg.get("projectRoot") or cfg.get("datasetRoot") or str(cwd)
    base = Path(base_str).expanduser().resolve()
    # If user points at ".../Aging", accept it
    if base.name.lower() == "aging":
        return base
    # If ".../CUI Workflows" and it contains "Aging", use it
    if (base / "Aging").is_dir():
        return base / "Aging"
    return base


def _compute_project_root(cfg: Dict[str, Any], project_root_arg: Optional[str], subject_root_arg: Optional[str], cwd: Path) -> Path:
    base_str = project_root_arg or cfg.get("projectRoot") or subject_root_arg or cfg.get("subjectRoot") or cfg.get("datasetRoot") or str(cwd)
    return Path(base_str).expanduser().resolve()


def _scan_max_subject_number(aging_root: Path) -> int:
    """
    Scan for subjectNNN folders under aging_root/<Sex>/<Ethnicity>/subjectNNN.
    """
    max_n = 0
    if not aging_root.is_dir():
        return 0
    pat = re.compile(r"^subject(\d+)$", re.IGNORECASE)

    for sex_dir in aging_root.iterdir():
        if not sex_dir.is_dir():
            continue
        for eth_dir in sex_dir.iterdir():
            if not eth_dir.is_dir():
                continue
            for subj_dir in eth_dir.iterdir():
                if not subj_dir.is_dir():
                    continue
                m = pat.match(subj_dir.name)
                if m:
                    try:
                        max_n = max(max_n, int(m.group(1)))
                    except Exception:
                        pass
    return max_n


def _locate_subject_folder(
    aging_root: Path,
    subject_folder_name: str,
    sex: Optional[str],
    ethnicity: Optional[str],
) -> Optional[Path]:
    # Fast path if sex+eth known
    if sex and ethnicity:
        cand = aging_root / _safe_folder(sex) / _safe_folder(ethnicity) / subject_folder_name
        if cand.is_dir():
            return cand

    # Otherwise scan
    pat = re.compile(r"^" + re.escape(subject_folder_name) + r"$", re.IGNORECASE)
    if aging_root.is_dir():
        for sex_dir in aging_root.iterdir():
            if not sex_dir.is_dir():
                continue
            for eth_dir in sex_dir.iterdir():
                if not eth_dir.is_dir():
                    continue
                cand = eth_dir / subject_folder_name
                if cand.is_dir() and pat.match(cand.name):
                    return cand
    return None


def _write_subject_notes_json(subject_folder: Path, subject_id: str, notes: str, meta: Dict[str, Any]) -> Path:
    subject_folder.mkdir(parents=True, exist_ok=True)
    out_path = subject_folder / "subject_notes.json"

    payload = {
        "schema_version": "a360_subject_notes_v1",
        "subject": subject_id,
        "notes": notes or "",
        "notes_meta": meta or {},
        "updated_utc": _utc_now_iso(),
    }

    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    return out_path


def _ensure_excel_exists(excel_path: Path) -> None:
    try:
        import openpyxl  # type: ignore
    except Exception:
        raise RuntimeError("openpyxl is not installed; cannot update Excel mirror")

    if excel_path.exists():
        return

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subjects"
    headers = ["SubjectID", "Base_Path", "Sex", "Ethnicity_Group", "Fitzpatrick_Tone", "Notes", "Image_Set_Status", "Last_Updated_Utc"]
    for j, h in enumerate(headers, 1):
        ws.cell(1, j).value = h

    wb.create_sheet("Images")
    wb.create_sheet("Prompts_Auto")
    wb.save(excel_path)


def _excel_upsert_subject(
    excel_path: Path,
    subject_id: str,
    base_path_rel: str,
    sex: str,
    ethnicity: str,
    fitz: str,
    notes: str,
) -> Tuple[bool, Optional[str]]:
    """
    Best-effort mirror update into Excel.
    Returns (updated, error).
    """
    if not excel_path:
        return False, None

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        return False, f"openpyxl not available: {e}"

    try:
        _ensure_excel_exists(excel_path)
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Subjects"] if "Subjects" in wb.sheetnames else wb.create_sheet("Subjects")

        required = ["SubjectID", "Base_Path", "Sex", "Ethnicity_Group", "Fitzpatrick_Tone", "Notes", "Image_Set_Status", "Last_Updated_Utc"]

        # Build / ensure header row
        header_vals = [c.value for c in ws[1]]
        if not any(v is not None and str(v).strip() for v in header_vals):
            for j, h in enumerate(required, 1):
                ws.cell(1, j).value = h
            header_vals = required[:]

        # Extend header if missing cols
        header = [str(v).strip() if v is not None else "" for v in header_vals]
        for col in required:
            if col not in header:
                header.append(col)
                ws.cell(1, len(header)).value = col

        col_idx = {name: header.index(name) + 1 for name in header if name}

        # Find existing row
        subj_col = col_idx.get("SubjectID", 1)
        row = None
        for r in range(2, ws.max_row + 1):
            v = ws.cell(r, subj_col).value
            if isinstance(v, str) and v.strip().upper() == subject_id.upper():
                row = r
                break

        if row is None:
            row = ws.max_row + 1

        ws.cell(row, col_idx["SubjectID"]).value = subject_id
        ws.cell(row, col_idx["Base_Path"]).value = base_path_rel
        ws.cell(row, col_idx["Sex"]).value = sex
        ws.cell(row, col_idx["Ethnicity_Group"]).value = ethnicity
        ws.cell(row, col_idx["Fitzpatrick_Tone"]).value = fitz
        ws.cell(row, col_idx["Notes"]).value = notes or ""
        if "Image_Set_Status" in col_idx and not ws.cell(row, col_idx["Image_Set_Status"]).value:
            ws.cell(row, col_idx["Image_Set_Status"]).value = "Not started"
        if "Last_Updated_Utc" in col_idx:
            ws.cell(row, col_idx["Last_Updated_Utc"]).value = _utc_now_iso()

        wb.save(excel_path)
        return True, None
    except Exception as e:
        return False, str(e)


class _QuietParser(argparse.ArgumentParser):
    def error(self, message: str) -> None:  # type: ignore[override]
        raise ValueError(message)


def _parse_json_meta(meta_str: Optional[str]) -> Dict[str, Any]:
    if not meta_str:
        return {}
    try:
        v = json.loads(meta_str)
        return v if isinstance(v, dict) else {}
    except Exception:
        return {}


# ----------------------------
# commands
# ----------------------------

def cmd_create_subject(argv: list[str], cwd: Path, cfg: Dict[str, Any]) -> int:
    p = _QuietParser(prog="update_subject_notes.py create-subject", add_help=True)
    p.add_argument("create_subject", help=argparse.SUPPRESS)
    p.add_argument("--sex", required=True)
    p.add_argument("--ethnicity", required=True)
    p.add_argument("--fitz", required=True)
    p.add_argument("--notes", default="")
    p.add_argument("--subjectRoot", default=None)
    p.add_argument("--projectRoot", default=None)
    p.add_argument("--excel", default=None)
    p.add_argument("--timelineFolderName", default=None)

    try:
        args = p.parse_args(argv)
    except ValueError as e:
        _eprint(str(e))
        return 2

    sex = _normalize_sex(args.sex)
    eth = _safe_folder(args.ethnicity, "Unsorted")
    fitz = str(args.fitz).strip() or "III"

    aging_root = _compute_aging_root(cfg, args.subjectRoot, args.projectRoot, cwd)
    project_root = _compute_project_root(cfg, args.projectRoot, args.subjectRoot, cwd)

    timeline_name = str(args.timelineFolderName or cfg.get("timelineFolderName") or "TimelineA").strip() or "TimelineA"

    # Allocate next subject
    max_n = _scan_max_subject_number(aging_root)
    next_n = max_n + 1
    width = max(3, len(str(next_n)))
    subject_folder_name = f"subject{next_n:0{width}d}"
    subject_id = f"S{next_n:0{width}d}"

    subject_folder = (aging_root / sex / eth / subject_folder_name).resolve()
    timeline_folder = (subject_folder / timeline_name).resolve()
    timeline_folder.mkdir(parents=True, exist_ok=True)

    base_rel = ""
    timeline_rel = ""
    try:
        base_rel = subject_folder.relative_to(aging_root).as_posix()
    except Exception:
        base_rel = subject_folder.name
    try:
        timeline_rel = timeline_folder.relative_to(aging_root).as_posix()
    except Exception:
        timeline_rel = (subject_folder.name + "/" + timeline_name)

    meta = {
        "sex": sex,
        "ethnicity_group": eth,
        "fitzpatrick_tone": fitz,
        "created_utc": _utc_now_iso(),
    }

    notes_path = _write_subject_notes_json(subject_folder, subject_id, args.notes or "", meta)

    excel_updated = False
    excel_error: Optional[str] = None
    excel_path = Path(args.excel).expanduser().resolve() if args.excel else None
    if excel_path:
        excel_updated, excel_error = _excel_upsert_subject(
            excel_path=excel_path,
            subject_id=subject_id,
            base_path_rel=base_rel,
            sex=sex,
            ethnicity=eth,
            fitz=fitz,
            notes=args.notes or "",
        )

    out = {
        "ok": True,
        "subjectId": subject_id,
        "sex": sex,
        "ethnicity": eth,
        "fitzpatrickTone": fitz,
        "notes": args.notes or "",
        "basePathRel": base_rel,
        "subjectFolderAbs": str(subject_folder),
        "timelineFolderAbs": str(timeline_folder),
        "timelineFolderRel": timeline_rel,
        "subjectNotesJsonPath": str(notes_path),
        "excelUpdated": bool(excel_updated),
    }
    if excel_error:
        out["excelError"] = excel_error

    print(json.dumps(out, ensure_ascii=False))
    return 0


def cmd_update_notes(argv: list[str], cwd: Path, cfg: Dict[str, Any]) -> int:
    p = _QuietParser(prog="update_subject_notes.py", add_help=True)
    p.add_argument("--subject", required=True)
    p.add_argument("--notes", required=True)
    p.add_argument("--meta", default=None)
    p.add_argument("--subjectRoot", default=None)
    p.add_argument("--projectRoot", default=None)
    p.add_argument("--excel", default=None)

    try:
        args = p.parse_args(argv)
    except ValueError as e:
        _eprint(str(e))
        return 2

    subject_id, _, subject_folder_name = _normalize_subject_id(args.subject)
    meta = _parse_json_meta(args.meta)

    # Pull placement hints from meta
    sex = meta.get("sex") or meta.get("Sex") or meta.get("gender") or meta.get("Gender")
    ethnicity = meta.get("ethnicity_group") or meta.get("Ethnicity_Group") or meta.get("ethnicity") or meta.get("Ethnicity")
    fitz = meta.get("fitzpatrick_tone") or meta.get("Fitzpatrick_Tone") or meta.get("fitz")

    if isinstance(sex, str):
        sex = _normalize_sex(sex)
    else:
        sex = None
    if isinstance(ethnicity, str):
        ethnicity = _safe_folder(ethnicity, "Unsorted")
    else:
        ethnicity = None
    if not isinstance(fitz, str):
        fitz = ""

    aging_root = _compute_aging_root(cfg, args.subjectRoot, args.projectRoot, cwd)
    project_root = _compute_project_root(cfg, args.projectRoot, args.subjectRoot, cwd)

    subject_folder = _locate_subject_folder(aging_root, subject_folder_name, sex, ethnicity)

    if subject_folder is None:
        # Create in hinted bucket or Unsorted
        sex_dir = sex or "Unsorted"
        eth_dir = ethnicity or "Unsorted"
        subject_folder = (aging_root / _safe_folder(sex_dir) / _safe_folder(eth_dir) / subject_folder_name).resolve()

    # Ensure meta includes canonical keys
    meta_out = dict(meta) if isinstance(meta, dict) else {}
    if sex:
        meta_out.setdefault("sex", sex)
    if ethnicity:
        meta_out.setdefault("ethnicity_group", ethnicity)
    if fitz:
        meta_out.setdefault("fitzpatrick_tone", fitz)

    notes_path = _write_subject_notes_json(subject_folder, subject_id, args.notes or "", meta_out)

    base_rel = ""
    try:
        base_rel = subject_folder.relative_to(aging_root).as_posix()
    except Exception:
        base_rel = subject_folder.name

    excel_updated = False
    excel_error: Optional[str] = None
    excel_path = Path(args.excel).expanduser().resolve() if args.excel else None
    if excel_path:
        excel_updated, excel_error = _excel_upsert_subject(
            excel_path=excel_path,
            subject_id=subject_id,
            base_path_rel=base_rel,
            sex=sex or "",
            ethnicity=ethnicity or "",
            fitz=str(fitz or ""),
            notes=args.notes or "",
        )

    out = {
        "ok": True,
        "subjectId": subject_id,
        "subjectFolderAbs": str(subject_folder),
        "basePathRel": base_rel,
        "subjectNotesJsonPath": str(notes_path),
        "excelUpdated": bool(excel_updated),
    }
    if excel_error:
        out["excelError"] = excel_error

    print(json.dumps(out, ensure_ascii=False))
    return 0


def main() -> int:
    cwd = Path.cwd()
    cfg = _load_config(cwd)

    argv = sys.argv[1:]

    if not argv:
        _eprint("No arguments provided")
        return 2

    if argv[0] == "create-subject":
        return cmd_create_subject(["create-subject"] + argv[1:], cwd, cfg)

    return cmd_update_notes(argv, cwd, cfg)


if __name__ == "__main__":
    raise SystemExit(main())
