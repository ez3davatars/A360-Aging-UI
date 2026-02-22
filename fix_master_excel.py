
import openpyxl
from pathlib import Path

excel_path = r"C:\Users\clone\OneDrive\Desktop\A360_AgingDataset_Master_Gemini_ComfyUI_v3_fullImages.xlsx"
wb = openpyxl.load_workbook(excel_path)
ws = wb["Subjects"]

cols = [cell.value for cell in ws[1]]
col_idx = {name: i + 1 for i, name in enumerate(cols)}

updated_count = 0
for row in range(2, ws.max_row + 1):
    sid = str(ws.cell(row, col_idx["SubjectID"]).value)
    base_path = str(ws.cell(row, col_idx["Base_Path"]).value)
    
    # Fix S003 metadata
    if sid == "S003":
        print(f"Correcting S003 metadata...")
        ws.cell(row, col_idx["Sex"]).value = "Female"
        ws.cell(row, col_idx["Ethnicity_Group"]).value = "Norwegian"
        ws.cell(row, col_idx["Folder_Name"]).value = "Norwegian_Female"
        base_path = "Female/Norwegian/subject003" # Force correct path
        ws.cell(row, col_idx["Base_Path"]).value = base_path
        updated_count += 1
        continue

    # Fix redundant Aging/ prefix for all
    if base_path.startswith("Aging/"):
        new_path = base_path.replace("Aging/", "", 1)
        print(f"Updating {sid}: {base_path} -> {new_path}")
        ws.cell(row, col_idx["Base_Path"]).value = new_path
        updated_count += 1
    elif base_path.startswith("Aging\\"):
        new_path = base_path.replace("Aging\\", "", 1)
        print(f"Updating {sid}: {base_path} -> {new_path}")
        ws.cell(row, col_idx["Base_Path"]).value = new_path
        updated_count += 1

wb.save(excel_path)
print(f"Done. Updated {updated_count} subjects.")
