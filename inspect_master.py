
import openpyxl
import os

path = r"C:\Users\clone\OneDrive\Desktop\A360_AgingDataset_Master_Gemini_ComfyUI_v3_fullImages.xlsx"
if not os.path.exists(path):
    print(f"File not found: {path}")
else:
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Subjects" in wb.sheetnames:
        ws = wb["Subjects"]
        cols = [cell.value for cell in ws[1]]
        with open("output_females.txt", "w", encoding="utf-8") as f:
            f.write(f"Columns: {cols}\n")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0: continue
                data = dict(zip(cols, row))
                if data.get("Sex") == "Female":
                    f.write(f"FOUND Female {data['SubjectID']} at row {i+1}:\n")
                    for k, v in data.items():
                        f.write(f"  {k}: {v}\n")
        print("Done. Results in output_females.txt")
    else:
        print(f"Sheet 'Subjects' not found in {wb.sheetnames}")
